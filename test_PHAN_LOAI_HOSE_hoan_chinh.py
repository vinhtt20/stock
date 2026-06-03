"""Unit tests cho PHAN_LOAI_HOSE_hoan_chinh — chạy hoàn toàn offline.

Mọi dữ liệu là DataFrame tổng hợp. Module import-được mà không gọi vnstock
vì việc import được hoãn vào main()/resolve_source().
"""
import os

import pandas as pd
import pytest

import PHAN_LOAI_HOSE_hoan_chinh as m


# --------------------------------------------------------------------------
# norm
# --------------------------------------------------------------------------
def test_norm_strips_accents_and_case():
    assert m.norm("Ngân Hàng") == "ngan hang"
    assert m.norm("  Bảo Hiểm ") == "bao hiem"


def test_norm_handles_none_and_nan():
    assert m.norm(None) == ""
    assert m.norm(float("nan")) == ""
    assert m.norm("nan") == ""


def test_norm_maps_d_with_stroke():
    # GIÁ TRỊ CAO NHẤT: nếu thiếu map đ->d, NFD giữ nguyên "đ" và regex
    # "bat dong san" không bao giờ khớp -> toàn bộ nhánh BĐS hỏng âm thầm.
    assert m.norm("Bất động sản") == "bat dong san"
    assert m.norm("Đầu tư") == "dau tu"


# --------------------------------------------------------------------------
# all_names — tha thứ cột vắng mặt
# --------------------------------------------------------------------------
def test_all_names_tolerates_absent_columns():
    row = {"icb_name_lv1": "Ngân Hàng"}  # thiếu lv2/lv3/lv4
    cols = ["icb_name_lv1", "icb_name_lv2", "icb_name_lv3", "icb_name_lv4"]
    out = m.all_names(row, cols)
    assert "ngan hang" in out
    # các cột vắng -> chuỗi rỗng, vẫn không lỗi
    assert out == "ngan hang |  |  | "


# --------------------------------------------------------------------------
# classify — đủ 6+ trường hợp
# --------------------------------------------------------------------------
COLS = ["icb_name_lv1", "icb_name_lv2", "icb_name_lv3", "icb_name_lv4"]


def test_classify_ngan_hang():
    row = {"icb_name_lv1": "Ngân hàng", "com_type_code": ""}
    assert m.classify(row, COLS) == m.LBL_NH


def test_classify_chung_khoan():
    row = {"icb_name_lv2": "Chứng khoán", "com_type_code": ""}
    assert m.classify(row, COLS) == m.LBL_CK


def test_classify_bao_hiem():
    row = {"icb_name_lv1": "Bảo hiểm", "com_type_code": ""}
    assert m.classify(row, COLS) == m.LBL_BH


def test_classify_real_estate_wins_over_financial():
    # vừa khớp BĐS vừa khớp KW tài chính -> BĐS thắng
    row = {
        "icb_name_lv1": "Bất động sản",
        "icb_name_lv2": "Dịch vụ tài chính",
        "com_type_code": "CK",
    }
    assert m.classify(row, COLS) == m.LBL_REAL_ESTATE


def test_classify_com_type_financial_with_blank_names():
    # com_type_code tài chính nhưng tên ICB trống -> vẫn tài chính, KHÔNG thiếu ICB
    for ctc, expected in [("NH", m.LBL_NH), ("CK", m.LBL_CK), ("BH", m.LBL_BH)]:
        row = {"icb_name_lv1": "", "com_type_code": ctc}
        assert m.classify(row, COLS) == expected


def test_classify_fin_other():
    # Chỉ có "Dịch vụ tài chính"/"tài chính", không NH/CK/BH, com_type_code phi TC
    row = {"icb_name_lv1": "Dịch vụ tài chính", "com_type_code": "DN"}
    assert m.classify(row, COLS) == m.LBL_FIN_OTHER


def test_classify_plain_non_financial():
    row = {"icb_name_lv1": "Công nghệ thông tin", "com_type_code": "DN"}
    assert m.classify(row, COLS) == m.LBL_NON_FIN


def test_classify_empty_name_cols_is_missing_icb():
    # name_cols=[] -> all_names rỗng, names_empty True -> thiếu ICB
    row = {"icb_name_lv1": "Ngân hàng", "com_type_code": ""}
    assert m.classify(row, []) == m.LBL_MISSING_ICB


def test_classify_missing_icb():
    row = {"icb_name_lv1": "", "icb_name_lv2": None, "com_type_code": ""}
    assert m.classify(row, COLS) == m.LBL_MISSING_ICB


# --------------------------------------------------------------------------
# build_dataset
# --------------------------------------------------------------------------
def _synthetic_hose():
    return pd.DataFrame(
        {
            "symbol": ["AAA", "BBB", "CCC", "DDD", "EEE", "FF", "GGG"],
            "organ_name": ["A Corp", "B Bank", "C Sec", "D RE", "E Tech",
                           "too short", "None"],
            "exchange": ["HOSE", "HSX", "HOSE", "HOSE", "HOSE", "HOSE", "HOSE"],
        }
    )


def _synthetic_icb():
    # dạng dài: mỗi symbol có icb_level 1 (đủ minh hoạ)
    rows = [
        ("AAA", "DN", 1, "0001", "Công nghệ"),
        ("BBB", "NH", 1, "8000", "Ngân hàng"),
        ("CCC", "CK", 1, "8700", "Chứng khoán"),
        ("DDD", "DN", 1, "8600", "Bất động sản"),
        ("EEE", "DN", 1, "9500", "Công nghệ thông tin"),
        # GGG: organ_name == "None" -> bị loại; không cần ICB
    ]
    return pd.DataFrame(
        rows,
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )


def test_build_dataset_merge_and_filters():
    df, name_cols = m.build_dataset(_synthetic_hose(), _synthetic_icb())
    syms = set(df["symbol"])
    # FF (2 ký tự) bị loại bởi bộ lọc 3 chữ cái; GGG bị loại do organ_name None
    assert "FF" not in syms
    assert "GGG" not in syms
    assert {"AAA", "BBB", "CCC", "DDD", "EEE"} <= syms
    assert any(c.startswith("icb_name_lv") for c in name_cols)


def test_build_dataset_missing_icb_handling():
    hose = pd.DataFrame(
        {
            "symbol": ["XXX"],
            "organ_name": ["Mystery Co"],
            "exchange": ["HOSE"],
        }
    )
    # XXX không có dòng ICB nào và com_type_code rỗng sau merge
    icb = pd.DataFrame(
        [("ZZZ", "DN", 1, "0001", "Công nghệ")],
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )
    df, _ = m.build_dataset(hose, icb)
    assert df.loc[df["symbol"] == "XXX", "phan_loai"].iloc[0] == m.LBL_MISSING_ICB


def test_build_dataset_label_correctness():
    hose = pd.DataFrame(
        {
            "symbol": ["AAA", "BBB", "CCC", "DDD", "EEE"],
            "organ_name": ["A Bank", "B Sec", "C RE", "D Tech", "E Co"],
            "exchange": ["HOSE", "HOSE", "HOSE", "HOSE", "HOSE"],
        }
    )
    icb = pd.DataFrame(
        [
            ("AAA", "NH", 1, "8000", "Ngân hàng"),
            ("BBB", "CK", 1, "8700", "Chứng khoán"),
            ("CCC", "DN", 1, "8600", "Bất động sản"),
            ("DDD", "DN", 1, "9500", "Công nghệ thông tin"),
            # EEE: không có dòng ICB -> Chưa phân loại
        ],
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )
    df, _ = m.build_dataset(hose, icb)
    by_sym = df.set_index("symbol")["phan_loai"].to_dict()
    assert by_sym["AAA"] == m.LBL_NH
    assert by_sym["BBB"] == m.LBL_CK
    assert by_sym["CCC"] == m.LBL_REAL_ESTATE
    assert by_sym["DDD"] == m.LBL_NON_FIN
    assert by_sym["EEE"] == m.LBL_MISSING_ICB
    # GIỮ kiểm tra đối soát số lượng (nhưng không phải khẳng định duy nhất).
    assert len(df) == int(df["phan_loai"].value_counts().sum())


def test_build_dataset_multilevel_pivot():
    hose = pd.DataFrame(
        {"symbol": ["AAA"], "organ_name": ["A Co"], "exchange": ["HOSE"]}
    )
    icb = pd.DataFrame(
        [
            ("AAA", "DN", 1, "1000", "Cấp 1"),
            ("AAA", "DN", 2, "2000", "Cấp 2"),
            ("AAA", "DN", 3, "3000", "Cấp 3"),
            ("AAA", "DN", 4, "4000", "Cấp 4"),
        ],
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )
    df, name_cols = m.build_dataset(hose, icb)
    for lv in (1, 2, 3, 4):
        assert f"icb_name_lv{lv}" in df.columns
        assert f"icb_code_lv{lv}" in df.columns
        assert df[f"icb_name_lv{lv}"].iloc[0] == f"Cấp {lv}"
        assert df[f"icb_code_lv{lv}"].iloc[0] == f"{lv}000"
    assert set(name_cols) == {f"icb_name_lv{lv}" for lv in (1, 2, 3, 4)}


# --------------------------------------------------------------------------
# _unique_sheet_name
# --------------------------------------------------------------------------
def test_unique_sheet_name_collision():
    used = {"cong_nghe"}
    out = m._unique_sheet_name("cong_nghe", used)
    assert out == "cong_nghe_2"
    used.add(out)
    out2 = m._unique_sheet_name("cong_nghe", used)
    assert out2 == "cong_nghe_3"


def test_unique_sheet_name_reserved():
    used = set(m.RESERVED_SHEETS)
    assert m._unique_sheet_name("Toan_bo", used) == "Toan_bo_2"
    assert m._unique_sheet_name("Phi_tai_chinh", used) == "Phi_tai_chinh_2"


def test_unique_sheet_name_truncates_to_31():
    long = "a" * 50
    out = m._unique_sheet_name(long, set())
    assert len(out) <= 31
    used = {out}
    out2 = m._unique_sheet_name(long, used)
    assert len(out2) <= 31
    assert out2.endswith("_2")


def test_unique_sheet_name_exact_truncation_value():
    # base >31 ký tự đã bị chiếm -> base[:31-len("_2")] + "_2" = base[:29] + "_2"
    long = "a" * 50
    truncated = long[:31]  # 31 ký tự đầu, đã dùng
    used = {truncated}
    out = m._unique_sheet_name(long, used)
    assert out == "a" * 29 + "_2"
    assert len(out) == 31


def test_unique_sheet_name_empty_and_none_fallback():
    assert m._unique_sheet_name("", set()) == "sheet"
    assert m._unique_sheet_name(None, set()) == "sheet"
    assert m._unique_sheet_name("   ", set()) == "sheet"


# --------------------------------------------------------------------------
# make_reports
# --------------------------------------------------------------------------
def test_make_reports_warning_when_missing_present():
    hose = pd.DataFrame(
        {
            "symbol": ["XXX", "YYY"],
            "organ_name": ["Mystery", "Tech Co"],
            "exchange": ["HOSE", "HOSE"],
        }
    )
    icb = pd.DataFrame(
        [("YYY", "DN", 1, "9500", "Công nghệ thông tin")],
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )
    df, _ = m.build_dataset(hose, icb)
    report = m.make_reports(df)
    assert report["warning"] is not None
    assert "XXX" in report["warning"]
    assert "XXX" in report["missing_symbols"]


def test_make_reports_no_warning_when_no_missing():
    df, _ = m.build_dataset(_synthetic_hose(), _synthetic_icb())
    report = m.make_reports(df)
    if not report["missing_symbols"]:
        assert report["warning"] is None
    # bất biến: có cảnh báo khi và chỉ khi missing > 0
    assert (report["warning"] is not None) == (len(report["missing_symbols"]) > 0)


def test_make_reports_real_estate_only_in_sample_when_flag():
    df, _ = m.build_dataset(_synthetic_hose(), _synthetic_icb())
    r_excl = m.make_reports(df, include_real_estate=False)
    r_incl = m.make_reports(df, include_real_estate=True)
    assert m.LBL_REAL_ESTATE not in set(r_excl["sample"]["phan_loai"])
    assert m.LBL_REAL_ESTATE in set(r_incl["sample"]["phan_loai"])
    # Bật cờ -> mẫu lớn hơn (thêm BĐS vào).
    assert len(r_incl["sample"]) > len(r_excl["sample"])


def test_make_reports_missing_excluded_and_sorted_all_listed():
    hose = pd.DataFrame(
        {
            "symbol": ["MMM", "AAA", "TTT"],
            "organ_name": ["M Co", "A Co", "T Tech"],
            "exchange": ["HOSE", "HOSE", "HOSE"],
        }
    )
    # MMM và AAA thiếu ICB; chỉ TTT có ICB phi tài chính.
    icb = pd.DataFrame(
        [("TTT", "DN", 1, "9500", "Công nghệ thông tin")],
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )
    df, _ = m.build_dataset(hose, icb)
    report = m.make_reports(df)
    # Thiếu ICB KHÔNG vào mẫu.
    assert m.LBL_MISSING_ICB not in set(report["sample"]["phan_loai"])
    # Liệt kê TẤT CẢ mã thiếu (>=2) và đã sắp xếp.
    assert report["missing_symbols"] == ["AAA", "MMM"]


# --------------------------------------------------------------------------
# export_excel — không crash khi mẫu rỗng
# --------------------------------------------------------------------------
def test_export_excel_empty_sample(tmp_path):
    hose = pd.DataFrame(
        {"symbol": ["BBB"], "organ_name": ["B Bank"], "exchange": ["HOSE"]}
    )
    icb = pd.DataFrame(
        [("BBB", "NH", 1, "8000", "Ngân hàng")],
        columns=["symbol", "com_type_code", "icb_level", "icb_code", "icb_name"],
    )
    df, _ = m.build_dataset(hose, icb)  # chỉ có 1 mã tài chính -> mẫu phi TC rỗng
    out = tmp_path / "out.xlsx"
    m.export_excel(df, str(out))
    assert os.path.exists(out)


# --------------------------------------------------------------------------
# KBS_TO_ICB — bảng ánh xạ (Phụ lục A)
# --------------------------------------------------------------------------
def test_kbs_to_icb_has_25_keys():
    assert len(m.KBS_TO_ICB) == 25


def test_kbs_to_icb_buckets_and_excluded_cap1():
    excluded = {2: "BH", 3: "RE", 5: "CK", 11: "NH", 29: "FIN_OTHER"}
    for code, bucket in excluded.items():
        name, b, icb1, approx = m.KBS_TO_ICB[code]
        assert b == bucket
        assert icb1 is None  # nhóm bị loại -> không gán ICB cấp 1
    # 20 mã phi tài chính -> bucket NONFIN và CÓ ICB cấp 1
    nonfin_codes = set(m.KBS_TO_ICB) - set(excluded)
    assert len(nonfin_codes) == 20
    for code in nonfin_codes:
        _, b, icb1, _ = m.KBS_TO_ICB[code]
        assert b == "NONFIN"
        assert icb1 is not None and icb1 != ""


def test_kbs_to_icb_approx_only_specific_codes():
    approx_codes = {c for c, v in m.KBS_TO_ICB.items() if v[3]}
    assert approx_codes == {1, 17, 26, 28}


def test_kbs_to_icb_spot_check_cap1():
    assert m.KBS_TO_ICB[6][2] == "Công nghệ"
    assert m.KBS_TO_ICB[8][2] == "Y tế"
    assert m.KBS_TO_ICB[22][2] == "Tiện ích cộng đồng"
    assert m.KBS_TO_ICB[10][2] == "Vật liệu cơ bản"


def test_gan_tay_has_three_keys():
    assert set(m.GAN_TAY) == {"ADG", "YEG", "CLC"}


# --------------------------------------------------------------------------
# build_dataset_kbs
# --------------------------------------------------------------------------
def test_build_dataset_kbs_mapping_override_and_missing():
    hose = pd.DataFrame(
        {
            "symbol": ["AAA", "BBB", "CCC", "DDD", "ADG", "EEE"],
            "organ_name": ["A Co", "B Bank", "C RE", "D Tech", "ADG Co", "E ??"],
            "exchange": ["HOSE"] * 6,
        }
    )
    # AAA=6 (Công nghệ), BBB=11 (Ngân hàng, loại), CCC=3 (BĐS, loại),
    # DDD=8 (Y tế), ADG=có code rác nhưng GAN_TAY override, EEE=999 không có.
    kbs = pd.DataFrame(
        {
            "symbol": ["AAA", "BBB", "CCC", "DDD", "ADG", "EEE"],
            "industry_code": [6, 11, 3, 8, 999, 999],
        }
    )
    df, name_cols = m.build_dataset_kbs(hose, kbs)
    assert name_cols == []
    by = df.set_index("symbol")

    # Phi tài chính map đúng ICB cấp 1.
    assert by.loc["AAA", "phan_loai"] == m.LBL_NON_FIN
    assert by.loc["AAA", "nganh_icb_cap1"] == "Công nghệ"
    assert by.loc["DDD", "nganh_icb_cap1"] == "Y tế"

    # Tài chính + BĐS bị loại khỏi mẫu phi tài chính.
    assert by.loc["BBB", "phan_loai"] == m.LBL_NH
    assert by.loc["CCC", "phan_loai"] == m.LBL_REAL_ESTATE
    sample = m._non_financial_sample(df, include_real_estate=False)
    assert "BBB" not in set(sample["symbol"])
    assert "CCC" not in set(sample["symbol"])

    # GAN_TAY override (mã 999 nhưng có trong GAN_TAY).
    assert by.loc["ADG", "phan_loai"] == m.LBL_NON_FIN
    assert by.loc["ADG", "nganh_icb_cap1"] == m.GAN_TAY["ADG"]

    # Mã không có KBS code và không trong GAN_TAY -> thiếu ICB.
    assert by.loc["EEE", "phan_loai"] == m.LBL_MISSING_ICB


def test_build_dataset_kbs_overridable_column_name():
    hose = pd.DataFrame(
        {"symbol": ["AAA"], "organ_name": ["A Co"], "exchange": ["HOSE"]}
    )
    kbs = pd.DataFrame({"symbol": ["AAA"], "nganh_id": [6]})
    df, _ = m.build_dataset_kbs(hose, kbs, industry_code_col="nganh_id")
    assert df.set_index("symbol").loc["AAA", "nganh_icb_cap1"] == "Công nghệ"


def test_build_dataset_kbs_count_reconciliation():
    # 7 mã: 2 nonfin, 1 RE, 3 fin (NH/CK/BH), 1 missing -> tổng khớp.
    hose = pd.DataFrame(
        {
            "symbol": ["AAA", "BBB", "CCC", "DDD", "EEE", "FFF", "GGG"],
            "organ_name": ["a", "b", "c", "d", "e", "f", "g"],
            "exchange": ["HOSE"] * 7,
        }
    )
    kbs = pd.DataFrame(
        {
            "symbol": ["AAA", "BBB", "CCC", "DDD", "EEE", "FFF", "GGG"],
            "industry_code": [6, 8, 3, 11, 5, 2, 999],  # nonfin,nonfin,RE,NH,CK,BH,?
        }
    )
    df, _ = m.build_dataset_kbs(hose, kbs)
    counts = df["phan_loai"].value_counts()
    nonfin = counts.get(m.LBL_NON_FIN, 0)
    re_ = counts.get(m.LBL_REAL_ESTATE, 0)
    fin = (counts.get(m.LBL_NH, 0) + counts.get(m.LBL_CK, 0)
           + counts.get(m.LBL_BH, 0) + counts.get(m.LBL_FIN_OTHER, 0))
    missing = counts.get(m.LBL_MISSING_ICB, 0)
    assert nonfin == 2 and re_ == 1 and fin == 3 and missing == 1
    assert nonfin + re_ + fin + missing == len(df)


# --------------------------------------------------------------------------
# resolve_source — fallback VCI -> KBS, KHÔNG cần vnstock thật
# --------------------------------------------------------------------------
class _FakeListing:
    """Listing giả: điều khiển hành vi từng nguồn để test fallback."""

    def __init__(self, source, vci_raises=False, vci_empty=False):
        self.source = source
        self._vci_raises = vci_raises
        self._vci_empty = vci_empty

    def symbols_by_exchange(self):
        if self.source == "vci" and self._vci_empty:
            return pd.DataFrame(
                columns=["symbol", "organ_name", "exchange"])
        return pd.DataFrame(
            {"symbol": ["AAA"], "organ_name": ["A Co"], "exchange": ["HOSE"]}
        )

    def symbols_by_industries(self):
        if self.source == "vci":
            if self._vci_raises:
                raise RuntimeError("403 Forbidden")
            if self._vci_empty:
                return pd.DataFrame()
            return pd.DataFrame(
                [("AAA", "DN", 1, "0001", "Công nghệ")],
                columns=["symbol", "com_type_code", "icb_level",
                         "icb_code", "icb_name"],
            )
        # KBS: một mã ngành nguyên / symbol
        return pd.DataFrame({"symbol": ["AAA"], "icb_code": [6]})


def test_resolve_source_uses_vci_when_valid():
    hose, data, name = m.resolve_source(
        listing_factory=lambda s: _FakeListing(s))
    assert name == "VCI"
    assert "com_type_code" in data.columns  # bảng ICB dạng dài


def test_resolve_source_falls_back_on_vci_error():
    hose, data, name = m.resolve_source(
        listing_factory=lambda s: _FakeListing(s, vci_raises=True))
    assert name == "KBS"
    assert "icb_code" in data.columns


def test_resolve_source_falls_back_on_vci_empty():
    hose, data, name = m.resolve_source(
        listing_factory=lambda s: _FakeListing(s, vci_empty=True))
    assert name == "KBS"


# --------------------------------------------------------------------------
# export_reference_workbook — 6 sheet
# --------------------------------------------------------------------------
def test_export_reference_workbook_six_sheets(tmp_path):
    from openpyxl import load_workbook

    out = tmp_path / "ref.xlsx"
    m.export_reference_workbook(str(out), "KBS", "30/06/2025", m.KBS_TO_ICB)
    assert os.path.exists(out)
    wb = load_workbook(str(out))
    expected = [
        "Co so phan loai", "Phan nganh ICB", "VSIC QD36-2025",
        "Doi chieu ICB-VSIC", "Tai lieu trich dan", "Anh xa KBS-ICB",
    ]
    assert wb.sheetnames == expected
    for name in wb.sheetnames:
        assert len(name) <= 31
